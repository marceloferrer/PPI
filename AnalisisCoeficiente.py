# Basic comparison on bonds, showing mean and moving average. Then exporting to excel.
from ppi_client.models.account_movements import AccountMovements
from ppi_client.ppi import PPI
from ppi_client.models.orders_filter import OrdersFilter
from ppi_client.models.order_budget import OrderBudget
from ppi_client.models.order_confirm import OrderConfirm
from ppi_client.models.disclaimer import Disclaimer
from ppi_client.models.search_instrument import SearchInstrument
from ppi_client.models.search_marketdata import SearchMarketData
from ppi_client.models.search_datemarketdata import SearchDateMarketData
from ppi_client.models.order import Order
from ppi_client.models.instrument import Instrument
from datetime import datetime, timedelta
import asyncio
import json
import traceback
from openpyxl import Workbook
import pandas as pd
import matplotlib.pyplot as plt
from scipy.stats import norm
import numpy as np

# Change sandbox variable to False to connect to production environment
ppi = PPI(sandbox=False)

def api_to_dataframe(api_response):
    df_marketdata = pd.DataFrame(api_response)
    df_marketdata["date"] = pd.to_datetime(df_marketdata["date"].str[0:10])
    df_marketdata = df_marketdata.set_index(['date'])

    return df_marketdata

def main():
    try:
        # Change login credential to connect to the API
        ppi.account.login('User', 'Password')
        nombre1 = "GD30" # input("Ingrese nombre del primer bono a comparar\n")
        nombre2 = "AL30" # input("Ingrese nombre del segundo bono a comparar\n")
        dias = "10" # input("Ingrese la cantidad de días requeridos\n")

        now = datetime.now()
        previous = datetime.today() - timedelta(days=int(dias))

        # Obtengo las cotizaciones historicas
        print(f"Realizando busqueda de {nombre1} desde el {previous.strftime('%Y-%m-%d')}")
        market_data1 = ppi.marketdata.search(SearchDateMarketData(nombre1, "Bonos", "A-48HS",
                                                                 previous.strftime('%Y-%m-%d'),
                                                                 now.strftime('%Y-%m-%d')))

        market_data2 = ppi.marketdata.search(SearchDateMarketData(nombre2, "Bonos", "A-48HS",
                                                                 previous.strftime('%Y-%m-%d'),
                                                                 now.strftime('%Y-%m-%d')))

        df_market_data = api_to_dataframe(market_data1)
        df_market_data2 = api_to_dataframe(market_data2)

        print("\nGenerando medias")
        df1 = df_market_data[["price"]].copy()
        df2 = df_market_data2[["price"]].copy()

        df = df1.subtract(df2, fill_value=0)

        promedio = df[["price"]].mean()
        media10 = df["price"].rolling(window=10).mean()
        media30 = df[["price"]].rolling(window=30).mean()
        media100 = df[["price"]].rolling(window=100).mean()

        strPrecio = "Precio actual %.4f" % df[["price"]].iloc[-1]
        strPromedio = "Promedio %.4f" % promedio
        strMedia10 = "Media 10 días %.4f" % media10.iloc[-1]
        strMedia30 = "Media 30 días %.4f" % media30.iloc[-1]
        strMedia100 = "Media 100 días %.4f" % media100.iloc[-1]

        print(strPrecio)
        print(strPromedio)
        print(strMedia10)
        print(strMedia30)
        print(strMedia100)

        df[strPrecio] = df[["price"]]
        df[strPromedio] = promedio
        df[strMedia10] = media10
        df[strMedia30] = media30
        df[strMedia100] = media100

        df[[strPrecio, strPromedio, strMedia10, strMedia30, strMedia100]].plot()

        plt.show(block=True)

        # Creo el excel
        print("\nExportando MarketData")
        workbook = Workbook()
        sheet = workbook.active

        # Completo los encabezados
        sheet.cell(row=1, column=1).value = "Fecha"
        sheet.cell(row=1, column=2).value = "Cotizacion"
        sheet.cell(row=1, column=3).value = "Volumen"
        sheet.cell(row=1, column=4).value = "Apertura"
        sheet.cell(row=1, column=5).value = "Minimo"
        sheet.cell(row=1, column=6).value = "Maximo"

        sheet.cell(row=2, column=8).value = "Promedio"
        sheet.cell(row=2, column=9).value = "%.4f" % promedio
        sheet.cell(row=3, column=8).value = "Media 10"
        sheet.cell(row=3, column=9).value = "%.4f" % media10.iloc[-1]
        sheet.cell(row=4, column=8).value = "Media 30"
        sheet.cell(row=4, column=9).value = "%.4f" % media30.iloc[-1]
        sheet.cell(row=5, column=8).value = "Media 100"
        sheet.cell(row=5, column=9).value = "%.4f" % media100.iloc[-1]

        r = 1
        # LLeno el excel
        for ins in market_data:
            r = r + 1
            sheet.cell(row=r, column=1).value = ins['date']
            sheet.cell(row=r, column=2).value = ins['price']
            sheet.cell(row=r, column=3).value = ins['volume']
            sheet.cell(row=r, column=4).value = ins['openingPrice']
            sheet.cell(row=r, column=5).value = ins['min']
            sheet.cell(row=r, column=6).value = ins['max']

        # Guardo el excel
        workbook.save(filename=f"{nombre} ({dias}) {now.strftime('%Y-%m-%d %H_%M')}.xlsx")

    except Exception as message:
        print(datetime.now())
        print(message)

    input('\nPress Enter to exit')


if __name__ == '__main__':
    main()



